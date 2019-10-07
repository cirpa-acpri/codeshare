# Fraser Hay - Conestoga College, Kitchener Ontario - fhay@conestogac.on.ca
# -----------------------------------------------------------------------------
# Short title: A script to sort files into folders based on a look-up table.
# -----------------------------------------------------------------------------
# Longer description:
# Basically, we recieved a bunch of PDF reports from the province, which had program codes attached, 
# but I needed to quickly upload them to a folder structure based on academic schools on our website.
# So I made a look-up table of all the file names (highlight all the files, CTRL+SHIFT+Right-click 
# on them > Copy as path > paste in Excel), isolated the program codes using formulas in Excel, then
# looked up those codes in a program table I had to return their school. So at the end of that process,
# I had a table of all the reports (paths), and all the schools. So I made folders corresponding to
# those school names, and ran the script, and it sorted all the files into those folders based on the
# look-up table. Fancy.

# This is just an example of something I made the other day. Probably wasted a bit more time
# on it than I should have, and probably re-invented the wheel in terms of there being easier ways
# of doing this (like with JSON), but now I have something to show for it!
# Related: https://automatetheboringstuff.com/chapter9/
# -----------------------------------------------------------------------------


import openpyxl # This is the module that interacts with Excel files. https://openpyxl.readthedocs.io/en/stable/api/openpyxl.html
table = [] # Define my environment variables. "table" is going to be an array (list) of rows ("item")
headers = [] # Array of the headers in the file
item = {} # A dictionary that corresponds to a row of data, using the headers as the keys
confused = {} # ... I don't know why, but I had to save things to this before it would work.

# ------------------------

file = openpyxl.load_workbook(r"C:\Users\fhay\Desktop\temp.xlsx") # Load the reference table.
sheet = file.active # Set the active sheet to as an object.
# Assumption is that first row has headers.
for table_col in range(1, sheet.max_column + 1): # For every column...
    headers.append(sheet.cell(row=1, column=table_col).value) # Write the headers to the array.

for table_row in range(2, sheet.max_row + 1): # For every row in the file...
    for header in headers: # For every header...
        data = sheet.cell(row=table_row, column=headers.index(header) + 1).value
        item[header] = data # Encode a header-data pairing. Eg: {"Gender": "Female"}
    confused = item.copy() # I have no idea why, but I had to copy the item into another dictionary...
    table.append(confused) # ... in order to save it to the "table" array...

# Time to move around the files. Import the Shell Utilities module.
import shutil # https://docs.python.org/3/library/shutil.html

for x in table:  # For every item in the "table" array...
    # Next two lines: Playing with the string, I had trouble getting it working the way I needed. This is basically me encoding an extra quote and then removing it.
    temp = (r"C:\Users\fhay\Desktop\New folder\"")
    temp = temp[:-1]
    temp = temp + x["School"] # Add "School" to the destination path of the file. Eg. "C:\Users\fhay\Desktop\New folder\" becomes "C:\Users\fhay\Desktop\New folder\Engineering"
    shutil.copy(x["Path"], temp) # Copy the source file to the destination path!
    
# Finis

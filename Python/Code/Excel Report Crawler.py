# Fraser Hay - Conestoga College, Kitchener Ontario - fhay@conestogac.on.ca
# -----------------------------------------------------------------------------
# Short title: A script that isolated the comments from about 4,500 individual teacher evlaulation Excel reports.
# -----------------------------------------------------------------------------
# Longer description:
# We found ourselves needing to provide a dataset of comments submitted by students in their
# evaluations of teaching and learning, but we didn't have a source dataset - just the reports
# themselves - all ~4,500 of them. Knowing Python could probably help, and armed with some ideas
# from the internet (https://automatetheboringstuff.com/chapter12/), I set about turning a ridiculous
# job into a computer science solution, as below. To do this, I had to cobble together a reference
# "hit list" of all the report files (highlight all the files, CTRL+SHIFT+Right-click # on them >
# Copy as path > paste in Excel). Then, since all my reports were exactly the same layout in Excel,
# I knew I could just open them up, can copy out whatever I found in a particular area. 
#
# Looping through all 4,534 Excel report took about 15 minutes. But it seems to have worked
# flawlessly. And doing this by hand would have been many days of insanity. Woo Python!
# -----------------------------------------------------------------------------

import time # Simple timer package - since this is going to be a big loop job, it'll be nice to keep track of how long it's been running.
import openpyxl # This is the module that interacts with Excel files. https://openpyxl.readthedocs.io/en/stable/api/openpyxl.html
Comments = [] # Define <Comments> as a list/array
SAT_List = [] # Define <SAT_List> as a list/array
counter = 0 # Housekeeping

# Load the file "hit-list" into an array, just to make it easier to play with later. I mean we could probably loop through it line by line
# a different way, but I just want to keep this simple.
# Note: the random "r" denotes "real string". If I just put it in a string, Python tends to double \'s.
ListFile = openpyxl.load_workbook(r"S:\Institutional_Research\SATs & Blue\Dashboard\SAT Files Listing.xlsx")
sheet = ListFile.active # Make <sheet> the active sheet (upon loading the file).
# Note: range() is a built-in function of Python. It basically allows you to specify a Min/Max for a FOR loop.
for SAT in range(1, sheet.max_row + 1):   # For each row in the file... Note: max_row is a (openpyxl) parameter of a sheet - specifying the last valid row on the sheet.
    SAT_List.append(sheet.cell(row=SAT, column=1).value) # This loop just loops through all the rows, and encodes the values into the "hit-list" array.
# Now <SAT_List[]> is an array including all entries from the Excel listing. 

start = time.time() # Time checkpoint

# Main sequence loop.
for SAT in SAT_List: # For every SAT in our list, do the following...
    counter += 1 # Housekeeping
    print("Processing: " + str(counter) + " / " + str(len(SAT_List)) + " (" + str(int((counter / len(SAT_List) * 100))) + "%) - Time: " + str(round(time.time()-start, 1)) + "s") # Counter display, since this takes awhile.
    workbook = openpyxl.load_workbook(SAT) # Open the file
    sheet = workbook["Summary"] # Select the "Summary" Sheet
    personInfo = sheet["A3"].value # Save A3 to a variable.
    sheet = workbook["Comments"] # Move on to Comments. Select the sheet.
    # Appending Comments: Loop through all the comments in the sheet
    for row in range(3, sheet.max_row + 1): # Start on row 3, where comments begin; and for each row in the sheet...
        if sheet["B" + str(row)].value: # This is a boolean check: is the B-cell of this row empty? This line is satisfied only if this is True - meaning a value is present. If so, we proceed. If not, we break off the loop in this file.
            course = sheet["B" + str(row)].value # <course> now equals the value of B<row> {Eg. if this is the first loop, we start at 3, so B3}.
            comment = sheet["C" + str(row)].value # Similar to above, <comment> is populated with the comment
            comment = comment.replace("_x000D_", "") # Get rid of some garbage that apparently is inherent in encoding carriage returns with openpyxl?
            Comments.append([personInfo, course, comment]) # Append the identifiers, course ID and comment to a new line (list) in <Comments>
        else:
            break # If the row has no comment, we're done with this file. Next!

print("\nFinal Encoding time: " + str(round(time.time()-start, 1)) + "s") # After all comments are read, display a final time for the run.
            
# Save data as a CSV
import csv # Import the relevant module. https://docs.python.org/3/library/csv.html
# Define the file, overwrite (if applicable), no characters needed to separate lines, utf-8 rich encoding, assign all this to the object <csv_file>
with open(r"S:\Institutional_Research\SATs & Blue\Dashboard\Comments.csv", "w", newline="", encoding="utf-8-sig") as csv_file:
    temp = csv.writer(csv_file) # <temp> will hold the <writer> function of the <csv> module we imported, and act on <csv_file>
    for value in Comments: # For every entity [ID/Course/Comment] in <Comments>
        temp.writerows([value]) # Write a new row to the file.
